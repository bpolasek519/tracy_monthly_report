from typing import Callable, List
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.worksheet.page import PageMargins
import src.constants as con
import src.dataframe_helper as dh
from src.excel_styles import apply_column_widths, set_style, header_settings, alternate_color_fill, \
    center_style_the_headers, last_row_style, style_center_cols


def create_llc_sheet(df: pd.DataFrame, wb: Workbook, llc_type: str, last_row_cols: List[str],
                     cols_to_exclude: List[str], month: str, year: str) -> None:
    # Create a DataFrame
    if llc_type == 'WIP':
        final_df = dh.create_wip_df(df=df, title='LLC USPS WIP', for_fs=False, filename='USPS')
    elif llc_type == 'Outstanding':
        final_df = dh.create_outstanding_df(df, title='LLC USPS Outstanding', filename='USPS')
    else:
        final_df = dh.create_paid_df(df, 'LLC USPS Paid', month=month, filename='USPS')

    # Create a new sheet within the Workbook
    ws = wb.create_sheet(title=f'LLC {llc_type}')

    generate_worksheet_print_settings(llc_type, month, ws, year, is_fs=False)

    # Applying styles to headers
    center_style_the_headers(final_df, ws)

    # Map column names to their indices and apply styles
    col_names = final_df.columns
    col_indices = {}
    for col_name, style in con.STYLE_MAPPINGS.items():
        if col_name in col_names:
            col_idx = col_names.get_loc(col_name)
            col_indices[col_name] = col_idx
            if col_name == 'Comment':
                set_style(df=final_df, ws=ws, col_name=col_name, style=style, idx=col_idx, wrapText=True)
            else:
                set_style(df=final_df, ws=ws, col_name=col_name, style=style, idx=col_idx)

    # Apply accounting style to the total sum row
    last_row = final_df.iloc[-1]
    last_row_style_cols = last_row_cols
    last_row_style(final_df, last_row, last_row_style_cols, ws)

    # Apply center alignment style to specific columns
    center_algined_cols = ['Type: \nJOC, HB', 'Contract', 'Proj. #', 'Prob. \n C/O #']
    style_center_cols(center_algined_cols, final_df, ws)

    # Apply general styling to the cells in the sheet
    excluded_columns = cols_to_exclude
    for row, row_data in enumerate(final_df.itertuples(index=False), start=2):
        if row <= len(final_df):
            for col, value in enumerate(row_data[:-1], start=1):
                if col not in [col_indices[col_name] + 1 for col_name in excluded_columns]:
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'))
        elif row == len(final_df) + 1:
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=value)

    # Apply column widths, header settings, and alternate row color fill
    apply_column_widths(final_df, ws)
    header_settings(final_df, ws)
    alternate_color_fill(final_df, ws)


def create_fs_sheet(usps_df: pd.DataFrame, hcde_df: pd.DataFrame, misc_df: pd.DataFrame, buyboard_df: pd.DataFrame,
                    pca_df: pd.DataFrame, friendswood_df: pd.DataFrame, wb: Workbook, month: str, year: str,
                    fs_type: str, last_row_columns: List[str], columns_to_exclude_from_generic_styles: List[str],
                    df_creation_func: Callable) -> None:
    """
    Create an FS Paid sheet in an OpenPyXL Workbook, combining data from multiple sources.

    This function creates an OpenPyXL sheet named 'FS Paid' within the given Workbook, combining data from multiple
    DataFrames and applying various styles.

    :param df_creation_func: The function used to create the dataframe. Can be paid, outstanding, or wip.
    :param columns_to_exclude_from_generic_styles: The list of columns that have special styling, so they aren't passed
    into the generic styling.
    :param last_row_columns: The list of columns in the last row.
    :param fs_type: Paid, Outstanding, WIP
    :param usps_df: USPS data DataFrame.
    :param hcde_df: HCDE data DataFrame.
    :param misc_df: Misc data DataFrame.
    :param buyboard_df: Buyboard data DataFrame.
    :param pca_df: PCA data DataFrame.
    :param friendswood_df: Friendswood data DataFrame.
    :param wb: The target Workbook where the sheet will be created.
    :param month: The month for which the data is being created.
    :param year: The year for which the data is being created.
    """

    # Retrieve and process data from various sources
    final_buyboard_df, final_friendswood_df, final_hcde_df, final_misc_df, final_pca_df, final_usps_df = (
        clean_col_names(buyboard_df, friendswood_df, hcde_df, misc_df, pca_df, usps_df, func=df_creation_func,
                        name=fs_type, month=month))

    # Create a blank row DataFrame with NaN values
    blank_row = pd.DataFrame({col: [np.nan] for col in final_usps_df.columns})

    # Merge the multiple dataframes and note where excluded rows are for styling. This includes the 'total' row,
    # the blank row, and the next header row.
    df_lists = [final_usps_df, final_misc_df, final_buyboard_df, final_hcde_df, final_pca_df, final_friendswood_df]
    final_df, style_skipped_rows, total_rows = retrieve_skipped_rows(blank_row, df_lists)

    total_df = final_df.iloc[total_rows]

    # Calculate and add the total amount of money
    sums = {col: total_df[col].sum() for col in last_row_columns}
    total_row = {'Type: \nJOC, HB': f'Total {fs_type} {month} {year}', **sums}
    total_row_df = pd.DataFrame(total_row, index=[0])
    final_df = pd.concat([final_df, blank_row, total_row_df], ignore_index=True)

    final_df.reset_index(drop=True, inplace=True)

    ws = wb.create_sheet(title=f'FS {fs_type}')

    generate_worksheet_print_settings(fs_type, month, ws, year, is_fs=True)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False

    ws.page_setup.margins = PageMargins(
        left=0.2,  # Left margin
        right=0.2,  # Right margin
        top=1,  # Top margin
        bottom=0.25,  # Bottom margin
        header=0.3,  # Header margin
        footer=0.3
    )

    ws.print_title_rows = '1:1'

    # Setting a specified height to all the cells
    ws.sheet_format.defaultRowHeight = con.SHEET_ROW_HEIGHT
    ws.sheet_format.customHeight = True

    print_header_text = f'Facilities Sources\n{fs_type.capitalize()} Report\n{month} {year}'
    ws.oddHeader.center.text = print_header_text
    ws.evenHeader.center.text = print_header_text

    # Centers the header row
    center_style_the_headers(final_df, ws)

    # Map column names to their indices and apply styles
    col_names = final_df.columns
    col_indices = {}
    for col_name, style in con.STYLE_MAPPINGS.items():
        if col_name in col_names:
            col_idx = col_names.get_loc(col_name)
            col_indices[col_name] = col_idx
            if col_name == 'Comment':
                set_style(df=final_df, ws=ws, col_name=col_name, style=style, idx=col_idx, wrapText=True)
            else:
                set_style(df=final_df, ws=ws, col_name=col_name, style=style, idx=col_idx)

    # Apply accounting style to the total sum rows
    for col in last_row_columns:
        col_idx = final_df.columns.get_loc(col)
        for row in style_skipped_rows[:-2]:
            value = final_df.iloc[row, col_idx]
            cell = ws.cell(row=row + 2, column=col_idx + 1, value=value)
            cell.style = con.ACCOUNTING_STYLE_NO_BORDER
        value = final_df.iloc[-1][col]
        cell = ws.cell(row=len(final_df) + 1, column=col_idx + 1, value=value)
        cell.style = con.ACCOUNTING_STYLE_NO_BORDER

    # Apply center alignment style to specific columns
    center_algined_cols = ['Type: \nJOC, HB', 'Contract', 'Proj. #', 'Prob. \n C/O #']
    style_center_cols(center_algined_cols, final_df, ws)

    # Need to shift the rows down 2 due to the way the Excel sheet locates them.
    shifted_style_skipped_rows = [x + 2 for x in style_skipped_rows]

    # Apply general styling to the cells in the sheet
    for row, row_data in enumerate(final_df.itertuples(index=False), start=2):
        if row not in shifted_style_skipped_rows:
            for col, value in enumerate(row_data, start=1):
                if col not in [col_indices[col_name] + 1 for col_name in columns_to_exclude_from_generic_styles]:
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'))
        else:
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=value)

    # Apply column widths, header settings, and alternate row color fill
    apply_column_widths(final_df, ws)
    header_settings(final_df, ws, shifted_style_skipped_rows)
    alternate_color_fill(final_df, ws, shifted_style_skipped_rows)


def generate_worksheet_print_settings(sheet_type, month, ws, year, is_fs):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False

    ws.page_setup.margins = PageMargins(
        left=0.2,  # Left margin
        right=0.2,  # Right margin
        top=1,  # Top margin
        bottom=0.25,  # Bottom margin
        header=0.3,  # Header margin
        footer=0.3
    )
    ws.print_title_rows = '1:1'

    ws.freeze_panes = 'H1'

    # Setting a specified height to all the cells
    ws.sheet_format.defaultRowHeight = con.SHEET_ROW_HEIGHT
    ws.sheet_format.customHeight = True

    if is_fs:
        if sheet_type == 'WIP':
            print_header_text = f'Facilities Sources\nWIP Report\n{month} {year}'
        else:
            print_header_text = f'Facilities Sources\n{sheet_type.capitalize()} Report\n{month} {year}'
    else:
        if sheet_type == 'WIP':
            print_header_text = f'Dura Pier Facility Service, LLC\nWIP Report\n{month} {year}'
        else:
            print_header_text = f'Dura Pier Facility Service, LLC\n{sheet_type.capitalize()} Report\n{month} {year}'
    ws.oddHeader.center.text = print_header_text
    ws.evenHeader.center.text = print_header_text


def clean_col_names(buyboard_df, friendswood_df, hcde_df, misc_df, pca_df, usps_df, func, name, month):
    usps_df.rename(columns={'Facility Name': 'Client', 'Address': 'Location', '%': 'Billed %'}, inplace=True)
    final_usps_df = func(usps_df, title=f'USPS {name}', month=month, filename='USPS')
    hcde_df.rename(columns={'Type: JOC/HB': 'Type:  JOC, CC, HB', 'Contract ': 'Contract', 'Billed $': 'Bill $'},
                   inplace=True)
    final_hcde_df = func(hcde_df, title=f'HCDE {name}', month=month, filename='HCDE')
    misc_col_names = misc_df.columns
    misc_df.rename(columns={'Type:  JOC, HB': 'Type:  JOC, CC, HB', misc_col_names[1]: 'Contract', 'Client ': 'Client',
                            'Billed $': 'Bill $', 'Comments': 'Comment'}, inplace=True)
    final_misc_df = func(misc_df, title=f'Misc. {name}', month=month, filename='Misc.')
    buyboard_df.rename(columns={'JOC/HB': 'Type:  JOC, CC, HB', 'Billed $': 'Bill $', 'Comments': 'Comment'},
                       inplace=True)
    final_buyboard_df = func(buyboard_df, title=f'Buyboard {name}', month=month, filename='Buyboard')
    pca_df.rename(columns={'JOC/HB': 'Type:  JOC, CC, HB', 'Contract #': 'Contract', 'Billed $': 'Bill $'},
                  inplace=True)
    final_pca_df = func(pca_df, title=f'PCA {name}', month=month, filename='PCA')
    friendswood_df.rename(columns={'JOC/HB': 'Type:  JOC, CC, HB', 'Contract #': 'Contract', 'Billed $': 'Bill $',
                                   'Billed       %': 'Billed %'}, inplace=True)
    final_friendswood_df = func(friendswood_df, title=f'Friendswood {name}', month=month, filename='Friendswood')
    return final_buyboard_df, final_friendswood_df, final_hcde_df, final_misc_df, final_pca_df, final_usps_df


def retrieve_skipped_rows(blank_row, df_lists):
    style_skipped_rows = []
    total_rows = []
    final_df = pd.DataFrame()
    for df in df_lists:
        if not df.empty:
            if final_df.empty:
                final_df = df
                style_skipped_rows.append(len(final_df.index) - 1)
                total_rows.append(len(final_df.index) - 1)
                style_skipped_rows.append(len(final_df.index))
                style_skipped_rows.append(len(final_df.index) + 1)
            else:
                final_df = pd.concat([final_df, blank_row, df], ignore_index=True)
                style_skipped_rows.append(len(final_df.index) - 1)
                total_rows.append(len(final_df.index) - 1)
                style_skipped_rows.append(len(final_df.index))
                style_skipped_rows.append(len(final_df.index) + 1)
        else:
            continue
    return final_df, style_skipped_rows, total_rows