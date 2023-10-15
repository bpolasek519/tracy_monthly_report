from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter

from src import constants as con


def apply_column_widths(final_df, ws):
    for header_name, width in con.COLUMN_WIDTHS.items():
        col_letter = None
        for col_idx, col in enumerate(final_df.columns, start=1):
            if col == header_name:
                col_letter = get_column_letter(col_idx)
                break
        if col_letter:
            ws.column_dimensions[col_letter].width = width


def set_style(df, ws, col_name, style, idx, wrapText: bool = False):
    for row, value in enumerate(df[col_name][:-1], start=2):
        cell = ws.cell(row=row, column=idx + 1, value=value)
        cell.style = style
        if wrapText:
            cell.alignment = cell.alignment.copy(wrapText=True)


def header_settings(df, ws, shifted_style_skipped_rows=None):
    # This applies the header color of light yellow
    if shifted_style_skipped_rows is None:
        shifted_style_skipped_rows = []

    header_fill_color = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill_color
        cell.font = Font(bold=True)

    if not shifted_style_skipped_rows:
        for cell in ws[len(df.index) + 1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')
    else:
        for value in shifted_style_skipped_rows:
            for cell in ws[value]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
                cell.border = Border()
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.border = Border()
        cell.alignment = Alignment(horizontal='left')


def alternate_color_fill(df, ws, shifted_style_skipped_rows=None):
    """
    This function alternates a fill color for every other row. It excludes the header and the final row. If the sheet is
    a merge of multiple dataframes, it also excludes all the total rows that are located using shifted_df_lengths.
    :param df: The final dataframe to be written to the Excel sheet
    :param ws: The Excel sheet.
    :param shifted_style_skipped_rows: The rows to exclude. OPTIONAL.
    """
    if shifted_style_skipped_rows is None:
        shifted_style_skipped_rows = []

    fill_color = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    for row_index, row in df.iloc[2:-1:2].iterrows():
        if not shifted_style_skipped_rows:
            for col_index, value in enumerate(row):
                cell = ws.cell(row=row_index + 2, column=col_index + 1)  # +2 to account for 0-based index and header
                cell.value = value
                cell.fill = fill_color
        else:
            if row_index + 2 in shifted_style_skipped_rows:
                for col_index, value in enumerate(row):
                    cell = ws.cell(row=row_index + 2, column=col_index + 1)
                    cell.value = value
            else:
                for col_index, value in enumerate(row):
                    cell = ws.cell(row=row_index + 2, column=col_index + 1)
                    cell.value = value
                    cell.fill = fill_color


def center_style_the_headers(final_df, ws):
    for col, col_name in enumerate(final_df.columns, start=1):
        cell = ws.cell(row=1, column=col, value=col_name)
        cell.style = con.CENTER_STYLE
        cell.alignment = cell.alignment.copy(wrapText=True)


def last_row_style(final_df, last_row, last_row_style_cols, ws):
    for col in last_row_style_cols:
        col_idx = final_df.columns.get_loc(col)
        value = last_row[col]
        cell = ws.cell(row=len(final_df) + 1, column=col_idx + 1, value=value)
        cell.style = con.ACCOUNTING_STYLE_NO_BORDER


def style_center_cols(center_algined_cols, final_df, ws):
    for col_name in center_algined_cols:
        col_idx = final_df.columns.get_loc(col_name) + 1
        for row, value in enumerate(final_df[col_name][:-1], start=2):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.style = con.CENTER_STYLE
